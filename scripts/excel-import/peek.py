#!/usr/bin/env python3
"""4ファイルの Excel お問合せ管理表のシート名・ヘッダー・先頭数行を表示。"""
from pathlib import Path
import openpyxl

DIR = Path(__file__).parent
FILES = sorted(DIR.glob("0*_*.xlsx"))

for fp in FILES:
    print(f"\n{'=' * 70}\n{fp.name}\n{'=' * 70}")
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: {sheet_name} ---")
        rows = list(ws.iter_rows(values_only=True, max_row=8))
        if not rows:
            print("(empty)")
            continue
        for i, r in enumerate(rows):
            cells = [("" if v is None else str(v))[:25] for v in r]
            # ファイル末尾の空セルは省く
            while cells and cells[-1] == "":
                cells.pop()
            print(f"  Row{i+1}: {cells}")
        print(f"  ... (total rows: {ws.max_row}, max_col: {ws.max_column})")
    wb.close()
