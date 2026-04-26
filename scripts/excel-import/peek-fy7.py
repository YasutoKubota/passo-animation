#!/usr/bin/env python3
"""令和7年度シートのサンプル行を見る。"""
from pathlib import Path
import openpyxl

DIR = Path(__file__).parent
FILES = sorted(DIR.glob("0*_*.xlsx"))


def normalize(s: str) -> str:
    """全角数字 → 半角、トリム"""
    if s is None:
        return ""
    s = str(s).strip()
    # 全角→半角の数字変換
    table = str.maketrans("０１２３４５６７８９", "0123456789")
    return s.translate(table)


for fp in FILES:
    print(f"\n{'=' * 70}\n{fp.name}\n{'=' * 70}")
    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
    target = None
    for sn in wb.sheetnames:
        if normalize(sn) == "令和7年度":
            target = sn
            break
    if not target:
        print(f"  → 令和7年度シートが見つかりません。シート一覧: {wb.sheetnames}")
        wb.close()
        continue

    ws = wb[target]
    print(f"  Sheet: {target!r}, max_row={ws.max_row}, max_col={ws.max_column}")
    rows = list(ws.iter_rows(values_only=True))
    print(f"  Header: {rows[0]}")
    print(f"\n  --- データ行（先頭 5 行 + 末尾 3 行） ---")
    data_rows = [r for r in rows[1:] if any(v not in (None, "") for v in r)]
    print(f"  非空データ行数: {len(data_rows)}")
    for i, r in enumerate(data_rows[:5]):
        print(f"  row {i+1}: {r}")
    if len(data_rows) > 5:
        print(f"  ...")
        for i, r in enumerate(data_rows[-3:]):
            print(f"  end-{2-i}: {r}")
    wb.close()
