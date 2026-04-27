#!/usr/bin/env python3
"""4 ファイルの「令和7年度」+「令和8年度」シートを読み、
intake_forms に挿入する SQL を出力する。

出力: scripts/excel-import/import-fy7.sql

使い方:
  python3 scripts/excel-import/import-fy7.py
  → SQL を確認してから:
  node scripts/run-migration.mjs scripts/excel-import/import-fy7.sql
"""
from __future__ import annotations
from pathlib import Path
import openpyxl
import datetime as dt
import json
import re
import uuid
import sys

DIR = Path(__file__).parent
OUT_SQL = DIR / "import-fy7.sql"
OUT_JSON_PREVIEW = DIR / "import-fy7-preview.json"

# 4 ファイル → studio_location マッピング
FILE_TO_STUDIO = {
    "01_駅前_お問合せ管理表.xlsx": "shushoku",   # 就職ゼミナール
    "02_創造空間_お問合せ管理表.xlsx": "sozo",
    "03_PAS_お問合せ管理表.xlsx": "pas_okazaki",
    "04_PAST_お問合せ管理表.xlsx": "pas_toyota",
}

# 取込対象の会計年度（令和7 = 2025/4-2026/3、令和8 = 2026/4-2027/3）
TARGET_FYS = [2025, 2026]
DATE_RANGE_START = dt.date(2025, 4, 1)
DATE_RANGE_END = dt.date(2027, 3, 31)
TODAY_YEAR = dt.date.today().year  # 2026

# 各 FY ごとのシート名の候補（全角・半角・末尾スペースを許容するため後で normalize）
def fy_to_sheet_name(fy):
    """会計年度 (西暦) → 探すべき和暦シート名（normalize 後の文字列）"""
    reiwa = fy - 2018
    return f"令和{reiwa}年度"


def normalize_sheet_name(s: str) -> str:
    """全角数字 → 半角、トリム"""
    if s is None:
        return ""
    s = str(s).strip()
    table = str.maketrans("０１２３４５６７８９", "0123456789")
    return s.translate(table)


def excel_to_date(value):
    """Excel のセル値を date に変換。datetime / int / float / str / None 対応。"""
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    if isinstance(value, (int, float)):
        # Excel serial date (1900 base)
        try:
            return (dt.datetime(1899, 12, 30) + dt.timedelta(days=int(value))).date()
        except OverflowError:
            return None
    if isinstance(value, str):
        s = value.strip()
        # 日付っぽい文字列は parse
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d", "%-m/%-d"):
            try:
                d = dt.datetime.strptime(s, fmt).date()
                return d
            except ValueError:
                continue
    return None


def cell_str(value) -> str:
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        return value.isoformat()
    return str(value).strip()


def map_source(route, route_note):
    """ルート文字列 → source_choices と詳細フィールド。

    分類軸（user 指示で「相談支援員と知人紹介は別物」「病院もリーフレット/紹介で別」）:
      newspaper       折込チラシ・新聞折込（ハピなびを含む）
      posting         ポスティング
      passerby        通りすがり
      homepage        HP・問合せフォーム・メール・LINE
      sns             Instagram・TikTok など
      hello_work      ハローワーク・若サポ
      city_office     市役所・保健所
      hospital_leaflet 病院（モノを見て）— 「病院においてあるリーフレット・パンフ」
      hospital_referral 病院（人から聞いて）— 「病院の先生・スタッフからの紹介」
      support_office  相談支援員からの紹介（事業所スタッフ）
      school          学校・特支（教師・先生からの紹介）
      internal        自社の他事業所スタッフ・利用者経由
      referral        個人紹介（友人・知人・宗教団体・近所など）
      other           その他（自社開催の利用相談会・「輪輪」など正体不明）
    """
    s = ((route or "") + " " + (route_note or "")).strip()
    detail = {}

    def has(*keywords):
        return any(k in s for k in keywords)

    # 1. 学校・特支 — 「先生」を含む場合、医療系の先生は別系統なので注意
    is_school_word = has("中学校", "高校", "小学校", "特支", "教諭", "学校")
    is_teacher_non_medical = (
        ("先生" in s)
        and not has("医", "クリニック", "病院", "心療内科", "メンタル")
    )
    if is_school_word or is_teacher_non_medical:
        detail["source_other"] = (route_note or route or "").strip()
        return ["school"], detail

    # 2. 病院系（医療機関）— リーフレット or 紹介を判定
    if has("心療内科", "クリニック", "病院", "医療機関", "メンタル"):
        is_leaflet = has("リーフレット", "パンフ", "チラシ", "設置", "置いて")
        is_via_person = has("先生", "スタッフ", "ナース", "看護", "医師", "より", "から") or "紹介" in s
        detail["source_hospital_name"] = (route_note or route or "").strip()
        if is_via_person and not is_leaflet:
            return ["hospital_referral"], detail
        # デフォルトはリーフレット（原典に「リーフレット設置分」など多い）
        return ["hospital_leaflet"], detail

    # 3. ハロワ / 若サポ（公的就労支援）
    if has("ハロワ", "HW", "ハローワーク", "若サポ", "ワカサポ", "サポステ"):
        return ["hello_work"], detail

    # 4. 市役所 / 保健所
    if has("市役所", "保健所"):
        return ["city_office"], detail

    # 5. 相談支援員からの紹介（「相談支援」キーワードがあれば優先）
    if has("相談支援", "相談員"):
        detail["source_facility_name"] = (route_note or route or "").strip()
        return ["support_office"], detail

    # 6. 自社の他事業所経由（PAS/PAST/創造空間/就職ゼミナール）
    is_internal_org = has("PAS", "PAST", "創造空間", "就職ゼミナール", "PASK", "駅前")
    is_internal_kw = has("元利用者", "出戻り") or s.strip() in (
        "PAS", "PAST", "創造空間", "就職ゼミナール",
        "/ 創造空間", "/ PAS", "/ PAST", "/ 就職ゼミナール",
    )
    if is_internal_org or is_internal_kw:
        detail["source_other"] = (route_note or route or "").strip()
        return ["internal"], detail

    # 7. 訪問看護 → 医療系の紹介として hospital_referral
    if has("訪問看護", "訪看"):
        detail["source_hospital_name"] = (route_note or route or "").strip()
        return ["hospital_referral"], detail

    # 8. 個人紹介（○○様・○○さん・知人・友人・紹介）
    has_sama_after_name = bool(re.search(r"[^\s／/]+様", s))
    has_san_after_name = bool(re.search(r"[^\s／/]+さん(より|から|紹介|$|\s)", s))
    if (
        has("様より", "様紹介", "様から", "から紹介", "より紹介",
            "の紹介", "ご紹介", "知人", "友人")
        or has_sama_after_name
        or has_san_after_name
        or "紹介" in s
    ):
        detail["source_other"] = (route_note or route or "").strip()
        return ["referral"], detail

    # 9. SNS
    if has("インスタ", "Instagram", "TikTok", "ティックトック", "SNS"):
        detail["source_sns_name"] = (route_note or route or "").strip()
        return ["sns"], detail

    # 10. ホームページ / メール / フォーム / LINE
    if has("ホームページ", "HP", "hp", "ＨＰ", "メール", "フォーム", "問合せフォーム", "LINE", "ライン"):
        return ["homepage"], detail

    # 11. 折込・リーフレット・チラシ・ハピなび
    if has("ハピなび", "ハピナビ", "折込", "折り込み", "折りこみ", "リーフレット", "チラシ"):
        return ["newspaper"], detail

    # 12. ポスティング
    if has("ポスティング"):
        return ["posting"], detail

    # 13. 通りがかり
    if has("通りがかり", "通りすがり"):
        return ["passerby"], detail

    # その他（自社開催イベント・正体不明など）
    if route or route_note:
        detail["source_other"] = (
            (route or "") + (" / " + route_note if route_note else "")
        ).strip()
    return ["other"], detail


def map_notebook(disability_type, notebook_present=""):
    """障がい種別文字列 → (notebook_status, notebook_grade)。

    手帳の有無 (file 04 のみ) もヒントに使う。
    'なし' / '未申請' なら 無、それ以外で「精神」「療育」「身体」「発達」を判定。
    """
    s = (disability_type or "") + " " + (notebook_present or "")
    if not s.strip():
        return ("", "")
    lower = s.lower()

    # 「なし」「未申請」「無し」「無」 = 手帳なし
    if any(k in s for k in ["なし", "未申請", "無し", "未取得", "申請中"]):
        if all(k not in s for k in ["精神", "療育", "身体", "発達"]):
            return ("無", "")

    grade = ""
    m = re.search(r"([1-6])\s*級", s)
    if m:
        grade = f"{m.group(1)}級"

    if "精神" in s or "発達" in s:
        return ("精神", grade)
    if "療育" in s or "知的" in s:
        return ("療育", grade)
    if "身体" in s:
        return ("身体", grade)
    return ("", grade)


def map_birth_date(age):
    """年齢 → birth_date (current_year - age)-01-01 の概算"""
    if age is None or age == "":
        return None
    try:
        a = int(age)
    except (ValueError, TypeError):
        return None
    if a < 1 or a > 99:
        return None
    return dt.date(TODAY_YEAR - a, 1, 1)


def safe_str(v):
    s = cell_str(v)
    return s if s else None


def find_col(header, name, alt_names=None):
    """ヘッダーに名前が出てくる位置を返す。"""
    targets = [name] + (alt_names or [])
    for idx, h in enumerate(header):
        h_str = cell_str(h)
        for t in targets:
            if h_str == t or h_str.startswith(t):
                return idx
    return None


def parse_row(row, columns):
    """行 → intake_forms カラム dict。対象 FY 範囲外 / 例 行は None を返す。"""
    no = cell_str(row[columns["No"]]) if columns.get("No") is not None else ""
    if no == "例":
        return None

    inquiry_date = excel_to_date(row[columns["お問合せ日"]]) if columns.get("お問合せ日") is not None else None
    if inquiry_date is None:
        return None
    if not (DATE_RANGE_START <= inquiry_date <= DATE_RANGE_END):
        return None

    name = cell_str(row[columns["名前"]]) if columns.get("名前") is not None else ""
    if not name:
        return None  # 名前ない行はスキップ

    route = safe_str(row[columns["ルート"]]) if columns.get("ルート") is not None else None
    route_note = safe_str(row[columns["ルート備考"]]) if columns.get("ルート備考") is not None else None
    source_choices, source_detail = map_source(route or "", route_note or "")

    gender = safe_str(row[columns["性別"]]) if columns.get("性別") is not None else None
    age = row[columns["年齢"]] if columns.get("年齢") is not None else None
    birth_date = map_birth_date(age)

    addr_main = safe_str(row[columns["住所"]]) if columns.get("住所") is not None else None
    town = safe_str(row[columns["町名"]]) if columns.get("町名") is not None else None
    address = " ".join([x for x in [addr_main, town] if x]) or None

    phone = safe_str(row[columns["電話"]]) if columns.get("電話") is not None else None
    email = safe_str(row[columns["メール"]]) if columns.get("メール") is not None else None
    disability_type = safe_str(row[columns["障がい種別"]]) if columns.get("障がい種別") is not None else None
    illness = safe_str(row[columns["病名"]]) if columns.get("病名") is not None else None
    notebook_present = safe_str(row[columns["手帳"]]) if columns.get("手帳") is not None else None
    notebook_status, notebook_grade = map_notebook(disability_type or "", notebook_present or "")

    support_office = safe_str(row[columns["相談支援事業所"]]) if columns.get("相談支援事業所") is not None else None
    support_office_contact = safe_str(row[columns["相談支援員"]]) if columns.get("相談支援員") is not None else None

    scheduled_visit_date = excel_to_date(row[columns["来所予定日"]]) if columns.get("来所予定日") is not None else None
    visit_date = excel_to_date(row[columns["見学"]]) if columns.get("見学") is not None else None
    trial_date = excel_to_date(row[columns["体験"]]) if columns.get("体験") is not None else None
    city_meeting_date = excel_to_date(row[columns["市役所面談"]]) if columns.get("市役所面談") is not None else None
    service_start = excel_to_date(row[columns["利用"]]) if columns.get("利用") is not None else None
    notes_raw = safe_str(row[columns["備考"]]) if columns.get("備考") is not None else None

    # 体験日が記録されていれば trial_sessions に1日分追加
    trial_sessions = []
    if trial_date:
        trial_sessions.append({"date": trial_date.isoformat(), "slot": "morning"})

    # 市役所面談は datetime にする（時刻は 12:00 とする）
    city_office_meeting_at = None
    if city_meeting_date:
        city_office_meeting_at = dt.datetime.combine(city_meeting_date, dt.time(12, 0)).isoformat()

    # メモ：細かい補助情報のみ staff_notes に追記
    # 「見学日」は専用カラム visited_at に保存するので staff_notes には書かない
    notes_parts = []
    if notes_raw:
        notes_parts.append(notes_raw)
    if email:
        notes_parts.append(f"メール: {email}")
    if support_office_contact:
        notes_parts.append(f"相談員: {support_office_contact}")
    if disability_type and not notebook_status:
        notes_parts.append(f"障がい種別: {disability_type}")
    staff_notes = "\n".join(notes_parts) if notes_parts else None

    # 電話番号文字列に「（母）」「(母)」のような所有者ヒントが含まれていたら抽出
    phone_owner = "self"
    phone_clean = phone
    if phone:
        m_owner = re.search(r"[（(]\s*(母|父|兄|弟|姉|妹|配偶者|妻|夫|祖母|祖父|本人|後見人|保護者)\s*[)）]", phone)
        if m_owner:
            owner_word = m_owner.group(1)
            phone_clean = re.sub(r"[（(]\s*[^()）（]+\s*[)）]", "", phone).strip()
            owner_map = {
                "本人": "self",
                "母": "mother",
                "父": "father",
                "兄": "sibling", "弟": "sibling", "姉": "sibling", "妹": "sibling",
                "配偶者": "spouse", "妻": "spouse", "夫": "spouse",
                "祖母": "guardian", "祖父": "guardian",
                "後見人": "guardian", "保護者": "guardian",
            }
            phone_owner = owner_map.get(owner_word, "other")

    rec = {
        "name": name,
        "furigana": "",  # 元データに無い
        "phone": phone_clean,
        "phone_owner": phone_owner,
        "birth_date": birth_date.isoformat() if birth_date else None,
        "gender": gender if gender in ("男", "女") else None,
        "address": address,
        "source_choices": source_choices,
        "source_facility_name": source_detail.get("source_facility_name"),
        "source_hospital_name": source_detail.get("source_hospital_name"),
        "source_sns_name": source_detail.get("source_sns_name"),
        "source_other": source_detail.get("source_other"),
        "experience_choices": [],
        "interested_work": [],
        "trial_sessions": trial_sessions,
        "illness_name": illness,
        "notebook_status": notebook_status if notebook_status else None,
        "notebook_grade": notebook_grade if notebook_grade else None,
        "support_office_name": support_office,
        "support_office_contact": support_office_contact,
        "inquiry_date": inquiry_date.isoformat(),
        "scheduled_visit_date": scheduled_visit_date.isoformat() if scheduled_visit_date else None,
        # Excel の「見学」列 = 実際に来所した日 → visited_at として保存
        "visited_at": visit_date.isoformat() if visit_date else None,
        "service_start_date": service_start.isoformat() if service_start else None,
        "city_office_meeting_at": city_office_meeting_at,
        "staff_notes": staff_notes,
        # submitted_at は 見学日 があればそれ、なければ inquiry_date
        "submitted_at": (visit_date or inquiry_date).isoformat(),
    }
    return rec


def find_columns(header):
    """ヘッダーから列インデックス辞書を作る。ファイルごとの差異を吸収。"""
    return {
        "No": find_col(header, "No.", ["No", "ＮＯ"]),
        "お問合せ日": find_col(header, "お問合せ日", ["問合せ日", "お問い合わせ日"]),
        "ルート": find_col(header, "ルート"),
        "ルート備考": find_col(header, "ルート（備考）", ["ルート備考"]),
        "名前": find_col(header, "名前（本人）", ["名前", "氏名", "お名前"]),
        "性別": find_col(header, "性別"),
        "年齢": find_col(header, "年齢"),
        "住所": find_col(header, "住所", ["市区町村", "市町村"]),
        "町名": find_col(header, "町名"),
        "電話": find_col(header, "電話番号", ["電話"]),
        "メール": find_col(header, "メールアドレス", ["メール"]),
        "障がい種別": find_col(header, "障がい種別", ["障害種別", "障害者種別"]),
        "病名": find_col(header, "病名"),
        "手帳": find_col(header, "手帳の有無", ["手帳"]),
        "相談支援事業所": find_col(header, "相談支援事業所"),
        "相談支援員": find_col(header, "相談支援員"),
        "来所予定日": find_col(header, "来所予定日"),
        "見学": find_col(header, "見学"),
        "体験": find_col(header, "体験"),
        "市役所面談": find_col(header, "市役所面談"),
        "利用": find_col(header, "利用"),
        "備考": find_col(header, "備考", ["最新の状況", "理由"]),
    }


def sql_escape(s):
    if s is None:
        return "NULL"
    if isinstance(s, list):
        # text[] リテラル: ARRAY['a', 'b']
        if len(s) == 0:
            return "ARRAY[]::text[]"
        return "ARRAY[" + ",".join(sql_escape(x) for x in s) + "]"
    if isinstance(s, dict) or (isinstance(s, list) and any(isinstance(x, dict) for x in s)):
        return "'" + json.dumps(s, ensure_ascii=False).replace("'", "''") + "'::jsonb"
    if isinstance(s, bool):
        return "TRUE" if s else "FALSE"
    if isinstance(s, (int, float)):
        return str(s)
    s = str(s)
    return "'" + s.replace("'", "''") + "'"


def main():
    all_records = []
    target_sheet_names = {fy_to_sheet_name(fy) for fy in TARGET_FYS}

    for fname, studio in FILE_TO_STUDIO.items():
        fp = DIR / fname
        if not fp.exists():
            print(f"WARN: ファイルがありません: {fp}", file=sys.stderr)
            continue
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)

        # 対象シートを全部洗い出す（令和7 / 令和8 両方）
        target_sheets = []
        for sn in wb.sheetnames:
            if normalize_sheet_name(sn) in target_sheet_names:
                target_sheets.append(sn)
        if not target_sheets:
            print(f"WARN: 対象シートなし ({target_sheet_names}): {fname}", file=sys.stderr)
            wb.close()
            continue

        for sheet in target_sheets:
            ws = wb[sheet]
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            cols = find_columns(header)

            count_in = 0
            count_out = 0
            for raw in rows[1:]:
                if not raw or all(v in (None, "") for v in raw):
                    continue
                count_in += 1
                rec = parse_row(raw, cols)
                if rec is None:
                    continue
                rec["studio_location"] = studio
                rec["_source_file"] = fname
                rec["_source_sheet"] = sheet
                all_records.append(rec)
                count_out += 1

            print(
                f"  {fname} / {sheet!r}: {count_out} / {count_in} 行を取込",
                file=sys.stderr,
            )
        wb.close()

    print(f"\n合計取込: {len(all_records)} 件\n", file=sys.stderr)

    # JSON プレビュー
    OUT_JSON_PREVIEW.write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    print(f"プレビュー: {OUT_JSON_PREVIEW}", file=sys.stderr)

    # SQL 出力
    lines = []
    lines.append("-- 令和7+8年度 (2025/4-2027/3) お問合せ管理表 取込")
    lines.append("-- 自動生成: scripts/excel-import/import-fy7.py")
    lines.append("-- 4 ファイル合計 {} 件\n".format(len(all_records)))
    lines.append("BEGIN;\n")
    # 冪等性: 既存の取込分を削除して再投入する。
    # サンプル (山田太郎 / 鈴木花子) と窪田さんの手入力テストデータは残す。
    lines.append(
        "DELETE FROM public.intake_forms\n"
        " WHERE submitted_at >= '2025-04-01'\n"
        "   AND submitted_at <  '2027-04-01'\n"
        "   AND id NOT IN (\n"
        "     '11111111-1111-1111-1111-111111111111',\n"
        "     '22222222-2222-2222-2222-222222222222',\n"
        "     '050e8173-c74b-4e20-9562-06cbc47342b3'\n"
        "   );\n"
    )

    for r in all_records:
        cols = []
        vals = []

        def add(col, v):
            cols.append(col)
            vals.append(sql_escape(v))

        add("studio_location", r["studio_location"])
        add("name", r["name"])
        add("furigana", r["furigana"])
        add("phone", r["phone"])
        add("phone_owner", r.get("phone_owner") or "self")
        add("birth_date", r["birth_date"])
        add("gender", r["gender"])
        add("address", r["address"])
        add("source_choices", r["source_choices"])
        if r.get("source_facility_name"):
            add("source_facility_name", r["source_facility_name"])
        if r.get("source_hospital_name"):
            add("source_hospital_name", r["source_hospital_name"])
        if r.get("source_sns_name"):
            add("source_sns_name", r["source_sns_name"])
        if r.get("source_other"):
            add("source_other", r["source_other"])
        add("experience_choices", r["experience_choices"])
        add("interested_work", r["interested_work"])
        # trial_sessions は jsonb literal が必要
        ts = r["trial_sessions"]
        cols.append("trial_sessions")
        vals.append(
            "'" + json.dumps(ts, ensure_ascii=False).replace("'", "''") + "'::jsonb"
        )
        add("illness_name", r["illness_name"])
        add("notebook_status", r["notebook_status"])
        add("notebook_grade", r["notebook_grade"])
        add("support_office_name", r["support_office_name"])
        add("support_office_contact", r["support_office_contact"])
        add("inquiry_date", r["inquiry_date"])
        add("scheduled_visit_date", r["scheduled_visit_date"])
        add("visited_at", r["visited_at"])
        add("service_start_date", r["service_start_date"])
        add("city_office_meeting_at", r["city_office_meeting_at"])
        add("staff_notes", r["staff_notes"])
        add("submitted_at", r["submitted_at"])

        sql = (
            "INSERT INTO public.intake_forms ("
            + ", ".join(cols)
            + ") VALUES ("
            + ", ".join(vals)
            + ");"
        )
        lines.append(sql)

    lines.append("\nCOMMIT;\n")
    OUT_SQL.write_text("\n".join(lines), encoding="utf-8")
    print(f"SQL 出力: {OUT_SQL}", file=sys.stderr)


if __name__ == "__main__":
    main()
